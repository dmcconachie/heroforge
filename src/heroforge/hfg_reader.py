"""
/tmp/hfg_reader.py — Read and convert HeroForge
.hfg character exports.

Usage:
  # Read from source workbook:
  uv run python /tmp/hfg_reader.py SRC <sheet> \
      [--rows R1:R2] [--cols C1:C2] [--nonempty]

  # Read from character .hfg export:
  uv run python /tmp/hfg_reader.py CHAR \
      [--rows R1:R2] [--cols C1:C2] [--nonempty]

  # Look up IDs:
  uv run python /tmp/hfg_reader.py lookup \
      --class-ids 10,21 --alignment-id 6

  # Search for text:
  uv run python /tmp/hfg_reader.py search \
      SRC|CHAR <sheet> <text>

  # Dump full character extract:
  uv run python /tmp/hfg_reader.py dump

  # Convert .hfg to .char.yaml:
  uv run python /tmp/hfg_reader.py convert \
      --race Human
"""

from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
)

SRC_PATH = Path(
    "/home/dmcconachie/coding/HeroForge-Anew/"
    "HeroForge Anew 3.5 v7.4.1.1-a0.xlsm"
)
CHAR_PATH = Path(
    "/home/dmcconachie/coding/heroforge/"
    "tests/integration/custom_characters/"
    "Drufus Occult Slayer (Lvl 22)"
    "_v7.4.1.2.a1(D&D 3.5).hfg"
)
CHAR_COPY = Path("/tmp/drufus.xlsx")

# ── alignment index → yaml key ─────────────────
# 1-based row offset into Stats & Character
# Details rows 35-43:
#   LG, NG, CG, LN, N, CN, LE, NE, CE
ALIGNMENT_MAP = {
    1: "lawful_good",
    2: "neutral_good",
    3: "chaotic_good",
    4: "lawful_neutral",
    5: "neutral",
    6: "chaotic_neutral",
    7: "lawful_evil",
    8: "neutral_evil",
    9: "chaotic_evil",
}

# ── ability index → key ───────────────────────
ABILITY_MAP = {
    1: "str",
    2: "dex",
    3: "con",
    4: "int",
    5: "wis",
    6: "cha",
}

# Row labels in ExportSheet D2:D7
ABILITY_ROWS = {
    2: "str",
    3: "dex",
    4: "con",
    5: "int",
    6: "wis",
    7: "cha",
}


# ── helpers ────────────────────────────────────


def _open_src():
    return openpyxl.load_workbook(
        SRC_PATH,
        data_only=True,
        keep_vba=True,
        read_only=True,
    )


def _open_char():
    if not CHAR_COPY.exists():
        shutil.copy(CHAR_PATH, CHAR_COPY)
    return openpyxl.load_workbook(CHAR_COPY, data_only=True)


def _parse_range(s: str) -> tuple[int, int]:
    a, b = s.split(":")
    return int(a), int(b)


def _parse_col_range(s: str) -> tuple[int, int]:
    a, b = s.split(":")
    return (
        column_index_from_string(a),
        column_index_from_string(b),
    )


def _clean_skill(name: str) -> str:
    """Strip footnote markers and fix sub-skills."""
    name = re.sub(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]", "", name)
    name = name.strip()
    return name


# Map raw HFG skill sub-names to proper YAML
# names. Skills like "Dance" (Perform sub-skill)
# need the parent prefix.
_SKILL_PARENT: dict[str, str] = {}
# Rows 109-114: Craft sub-skills
# Row 143+: Knowledge custom sub-skills
# Rows 155+: Perform sub-skills
# Rows 161-165: Profession sub-skills
_PERFORM_ROWS = range(155, 160)
_CRAFT_ROWS = range(109, 115)
_PROFESSION_ROWS = range(161, 166)
_KNOWLEDGE_CUSTOM_ROWS = range(143, 150)


# ── lookup tables from source workbook ─────────


def _load_class_map() -> dict[int, str]:
    """Class ID → name from SRC Class Info."""
    wb = _open_src()
    ws = wb["Class Info"]
    m: dict[int, str] = {}
    for r in range(4, 300):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None or b is None:
            continue
        aid = int(a)
        name = str(b).strip()
        # Skip separator rows
        if name.startswith("–") or name.startswith("—"):
            continue
        if name.startswith("Select"):
            continue
        m[aid] = name
    return m


def _load_prc_map() -> dict[int, str]:
    """
    PrC index (from Z column) → name.
    The Z column stores an offset into the PrC
    section of Class Info: PrC index N maps to
    Class Info row (N + 104).
    """
    wb = _open_src()
    ws = wb["Class Info"]
    m: dict[int, str] = {}
    for r in range(105, 500):
        b = ws.cell(row=r, column=2).value
        if b is None:
            continue
        name = str(b).strip()
        if name.startswith("–") or name.startswith("—"):
            continue
        if name.startswith("Select"):
            continue
        if name.startswith("Prestige"):
            continue
        idx = r - 104
        m[idx] = name
    return m


def _load_weapon_map() -> dict[int, str]:
    """Weapon ID → name from CW&A (offset 31)."""
    wb = _open_src()
    ws = wb["Class Weapons & Armor"]
    m: dict[int, str] = {}
    for r in range(33, 400):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        wid = r - 31
        name = str(a).rstrip(" ●").strip()
        m[wid] = name
    return m


def _load_deity_map() -> dict[int, str]:
    """Deity ID → name from Deities sheet."""
    wb = _open_src()
    ws = wb["Deities"]
    m: dict[int, str] = {}
    for r in range(3, 600):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        did = r - 2
        name = str(a).strip()
        m[did] = name
    return m


# ── character extraction ───────────────────────


def extract_char() -> dict:
    """
    Pull all structured data from the .hfg
    ExportSheet into a plain dict.
    """
    wb = _open_char()
    ws = wb["ExportSheet"]

    data: dict = {}

    # ── identity ───────────────────────────
    data["name"] = ws["E12"].value or ""
    data["alignment_id"] = ws["E10"].value
    data["deity_id"] = ws["E11"].value

    # ── ability scores (base) ──────────────
    data["ability_scores"] = {}
    for row, key in ABILITY_ROWS.items():
        v = ws.cell(row=row, column=5).value
        data["ability_scores"][key] = int(v) if v else 10

    # ── stat bumps (L9:L13 → lvls 4,8,12,16,20)
    bumps: list[tuple[int, str]] = []
    bump_levels = [4, 8, 12, 16, 20]
    for i, blvl in enumerate(bump_levels):
        v = ws.cell(row=9 + i, column=12).value
        if v and isinstance(v, (int, float)):
            bumps.append((blvl, ABILITY_MAP.get(int(v), "?")))
    data["stat_bumps"] = bumps

    # ── levels / classes ───────────────────
    # P(16)=char level, Q(17)=class ID,
    # S(19)=HP roll
    levels: list[dict] = []
    for r in range(9, 69):  # up to 60 levels
        lvl = ws.cell(row=r, column=16).value
        cid = ws.cell(row=r, column=17).value
        if lvl is None or cid is None:
            continue
        if isinstance(cid, (int, float)) and (int(cid) <= 1):
            if r > 9 + 30:
                break
            continue
        hp = ws.cell(row=r, column=19).value

        entry = {
            "level": int(lvl),
            "class_id": int(cid),
        }
        if hp == "Max":
            entry["hp_roll"] = "max"
        elif hp and isinstance(hp, (int, float)):
            entry["hp_roll"] = int(hp)
        else:
            entry["hp_roll"] = "max"

        levels.append(entry)
    data["levels"] = levels

    # ── skills per level ───────────────────
    # Rows 101-180, D=name, G(col7)=lvl1 ...
    skills: dict[int, dict[str, int]] = {}
    for r in range(101, 181):
        raw = ws.cell(row=r, column=4).value
        if raw is None:
            continue
        name = _clean_skill(str(raw))
        # Fix sub-skill naming
        if r in _PERFORM_ROWS:
            name = f"Perform ({name})"
        elif r in _CRAFT_ROWS:
            name = f"Craft ({name})"
        elif r in _PROFESSION_ROWS:
            name = f"Profession ({name})"
        elif r in _KNOWLEDGE_CUSTOM_ROWS:
            # "local: IUZ" → "Knowledge (local)"
            base = name.split(":")[0].strip()
            name = f"Knowledge ({base})"
        for lvl in range(1, 61):
            col = 6 + lvl  # G=7 is level 1
            v = ws.cell(row=r, column=col).value
            if v and isinstance(v, (int, float)) and v > 0:
                skills.setdefault(lvl, {})[name] = int(v)
    data["skills"] = skills

    # ── feats ──────────────────────────────
    # BY(77)=name, BZ(78)=selected, CA(79)=bonus
    # CB(80)=sub-list
    feats: list[dict] = []
    last_parent: str | None = None
    for r in range(3, 3700):
        name = ws.cell(row=r, column=77).value
        sel = ws.cell(row=r, column=78).value
        bonus = ws.cell(row=r, column=79).value
        cb = ws.cell(row=r, column=80).value

        if name and isinstance(name, str):
            last_parent = name

        is_selected = sel is True
        is_bonus = bonus is True

        if not (is_selected or is_bonus):
            continue

        feat: dict = {}
        if name and isinstance(name, str):
            feat["name"] = name
        elif last_parent:
            # Sub-selection row
            feat["name"] = last_parent
            if cb and isinstance(cb, (int, float)):
                feat["parameter_id"] = int(cb)

        feat["is_bonus"] = is_bonus
        if feat.get("name"):
            feats.append(feat)

    data["feats"] = feats

    # ── weapons ────────────────────────────
    weapons: list[dict] = []
    # Weapon slots: E25, G25, I25, K25, M25, O25
    weapon_cols = [5, 7, 9, 11, 13, 15]
    for col in weapon_cols:
        wid = ws.cell(row=25, column=col).value
        if wid and isinstance(wid, (int, float)) and int(wid) > 1:
            enh = ws.cell(row=26, column=col).value
            weapons.append(
                {
                    "weapon_id": int(wid),
                    "enhancement": (
                        int(enh) if enh and isinstance(enh, (int, float)) else 0
                    ),
                }
            )
    data["weapons"] = weapons

    return data


def dump_char():
    """Print extracted character data."""
    data = extract_char()
    class_map = _load_class_map()
    prc_map = _load_prc_map()
    weapon_map = _load_weapon_map()
    deity_map = _load_deity_map()

    align_id = data["alignment_id"]
    deity_id = data["deity_id"]

    print(f"Name: {data['name']}")
    print(f"Alignment: {align_id} -> {ALIGNMENT_MAP.get(align_id, '?')}")
    deity = deity_map.get(deity_id, "?")
    if deity.startswith("Select"):
        deity = "(none)"
    print(f"Deity: {deity_id} -> {deity}")
    print(f"Ability scores: {data['ability_scores']}")
    print(f"Stat bumps: {[(l, a) for l, a in data['stat_bumps']]}")

    print("\n--- Levels ---")
    for lvl in data["levels"]:
        cname = class_map.get(
            lvl["class_id"],
            f"?{lvl['class_id']}",
        )
        skills = data["skills"].get(lvl["level"], {})
        sk_str = (
            ", ".join(f"{k}:{v}" for k, v in skills.items())
            if skills
            else "(none)"
        )
        print(
            f"  Lvl {lvl['level']:2d}:"
            f" {cname:20s}"
            f" HP={lvl['hp_roll']!s:4s}"
            f"  Skills: {sk_str}"
        )

    print("\n--- Feats ---")
    for f in data["feats"]:
        src = "bonus" if f["is_bonus"] else "char"
        param = ""
        if "parameter_id" in f:
            param = f" (param={f['parameter_id']})"
        print(f"  [{src:5s}] {f['name']}{param}")

    print("\n--- Weapons ---")
    for w in data["weapons"]:
        wname = weapon_map.get(
            w["weapon_id"],
            f"?{w['weapon_id']}",
        )
        print(f"  {wname} (+{w['enhancement']} enh)")


# ── conversion to .char.yaml ──────────────────


def _infer_feat_levels(
    levels: list[dict],
    feats: list[dict],
    class_map: dict[int, str],
) -> dict[int, list[dict]]:
    """
    Best-effort assignment of feats to levels.

    D&D 3.5e rules:
    - Character feats at 1,3,6,9,12,15,18,21,24…
    - Human bonus feat at level 1
    - Fighter bonus feats at class lvls
      1,2,4,6,8,10,12,14,16,18,20

    Returns {character_level: [feat_entries]}.
    """
    char_feats = [f for f in feats if not f["is_bonus"]]
    bonus_feats = [f for f in feats if f["is_bonus"]]

    # Char feat slots: every 3 levels
    # + human bonus at level 1 (added as a
    # separate entry with source human_bonus)
    char_lvls = [
        1,
        3,
        6,
        9,
        12,
        15,
        18,
        21,
        24,
        27,
        30,
    ]
    # TODO: detect race for human bonus feat.
    # For now, add a human bonus slot at level 1.
    # The first feat goes as "character", second
    # at level 1 goes as "human_bonus".
    human_bonus_slot = True

    # Fighter bonus feat levels (class level)
    ftr_bonus_class_lvls = {
        1,
        2,
        4,
        6,
        8,
        10,
        12,
        14,
        16,
        18,
        20,
    }
    fighter_count = 0
    fighter_bonus_char_lvls: list[int] = []
    for lvl in levels:
        cname = class_map.get(lvl["class_id"], "")
        if cname == "Fighter":
            fighter_count += 1
            if fighter_count in ftr_bonus_class_lvls:
                fighter_bonus_char_lvls.append(lvl["level"])

    result: dict[int, list[dict]] = {}

    # Assign character feats
    ci = 0
    used_human_bonus = False
    for f in char_feats:
        if human_bonus_slot and not used_human_bonus and ci == 1:
            # Second char feat at level 1 is
            # the human bonus feat
            result.setdefault(1, []).append(
                {
                    "name": f["name"],
                    "source": "human_bonus",
                }
            )
            used_human_bonus = True
            continue
        if ci < len(char_lvls):
            lvl = char_lvls[ci]
            ci += 1
        else:
            lvl = -1  # unassigned
        result.setdefault(lvl, []).append(
            {
                "name": f["name"],
                "source": "character",
            }
        )

    # Assign bonus feats
    bi = 0
    for f in bonus_feats:
        entry: dict = {"name": f["name"]}
        if "parameter_id" in f:
            # Resolve weapon focus parameter
            entry["parameter_id"] = f["parameter_id"]

        if bi < len(fighter_bonus_char_lvls):
            lvl = fighter_bonus_char_lvls[bi]
            entry["source"] = "fighter_bonus"
            bi += 1
        else:
            lvl = -1
            entry["source"] = "bonus"

        result.setdefault(lvl, []).append(entry)

    return result


def convert_char(
    race: str,
    out_path: str,
) -> None:
    """
    Build a Character from .hfg data and save
    it via the engine's own save_character().
    """
    from heroforge.engine.character import (
        Character,
        CharacterLevel,
    )
    from heroforge.engine.persistence import (
        save_character,
    )
    from heroforge.ui.app_state import AppState

    data = extract_char()
    class_map = _load_class_map()
    weapon_map = _load_weapon_map()
    deity_map = _load_deity_map()

    app_state = AppState()
    app_state.load_rules()

    c = Character()
    c._class_registry_ref = app_state.class_registry
    c._feat_registry_ref = app_state.feat_registry

    # Identity
    c.name = data["name"]
    c.player = ""
    c.alignment = ALIGNMENT_MAP.get(data["alignment_id"], "neutral")
    deity_id = data["deity_id"]
    deity = deity_map.get(deity_id, "")
    if deity.startswith("Select"):
        deity = ""
    c.deity = deity

    # Ability scores
    for ab, val in data["ability_scores"].items():
        c.set_ability_score(ab, val)

    # Race
    from heroforge.engine.races import apply_race

    race_defn = app_state.race_registry.get(race)
    if race_defn:
        apply_race(race_defn, c)
    else:
        c.race = race

    # Infer feat-to-level assignment
    feat_map = _infer_feat_levels(
        data["levels"],
        data["feats"],
        class_map,
    )

    # Levels
    for lvl in data["levels"]:
        cid = lvl["class_id"]
        cname = class_map.get(cid, f"Unknown_{cid}")

        hp = lvl["hp_roll"]
        if hp == "max":
            # Look up hit die from class registry
            cls_def = app_state.class_registry.get(cname)
            if cls_def:
                hp = cls_def.hit_die
            else:
                hp = 8

        skills = data["skills"].get(lvl["level"], {})
        feats = feat_map.get(lvl["level"], [])

        c.levels.append(
            CharacterLevel(
                character_level=lvl["level"],
                class_name=cname,
                hp_roll=hp,
                skill_ranks=dict(sorted(skills.items())),
                feats=feats,
            )
        )

    if c.levels:
        c._invalidate_class_stats()

    # Equipment
    weapons = []
    for w in data["weapons"]:
        wname = weapon_map.get(
            w["weapon_id"],
            f"Unknown_{w['weapon_id']}",
        )
        weapons.append({"name": wname, "base": wname})
    c.equipment = {
        "armor": {},
        "shield": {},
        "worn": [],
        "weapons": weapons,
    }

    save_character(c, out_path)
    print(f"Saved to {out_path}")


# ── CLI subcommands ────────────────────────────


def cmd_read(args):
    if args.target == "SRC":
        wb = _open_src()
    else:
        wb = _open_char()

    if args.sheet == "?":
        print("Sheets:", wb.sheetnames)
        return

    ws = wb[args.sheet]
    r1 = 1
    r2 = min(ws.max_row or 1, 50)
    c1 = 1
    c2 = min(ws.max_column or 1, 20)

    if args.rows:
        r1, r2 = _parse_range(args.rows)
    if args.cols:
        c1, c2 = _parse_col_range(args.cols)

    for r in range(r1, r2 + 1):
        vals = []
        for c in range(c1, c2 + 1):
            v = ws.cell(row=r, column=c).value
            if args.nonempty and v is None:
                continue
            cl = get_column_letter(c)
            vals.append(f"{cl}{r}={v!r}")
        if vals:
            print(f"  Row {r}: {', '.join(vals)}")


def cmd_lookup(args):
    if args.class_ids:
        m = _load_class_map()
        ids = [int(x) for x in args.class_ids.split(",")]
        for cid in ids:
            print(f"  Class {cid}: {m.get(cid, 'UNKNOWN')}")

    if args.prc_ids:
        m = _load_prc_map()
        ids = [int(x) for x in args.prc_ids.split(",")]
        for pid in ids:
            print(f"  PrC {pid}: {m.get(pid, 'UNKNOWN')}")

    if args.alignment_id:
        aid = int(args.alignment_id)
        print(f"  Alignment {aid}: {ALIGNMENT_MAP.get(aid, 'UNKNOWN')}")

    if args.deity_id:
        m = _load_deity_map()
        did = int(args.deity_id)
        print(f"  Deity {did}: {m.get(did, 'UNKNOWN')}")

    if args.weapon_ids:
        m = _load_weapon_map()
        ids = [int(x) for x in args.weapon_ids.split(",")]
        for wid in ids:
            print(f"  Weapon {wid}: {m.get(wid, 'UNKNOWN')}")


def cmd_search(args):
    if args.target == "SRC":
        wb = _open_src()
    else:
        wb = _open_char()

    ws = wb[args.sheet]
    text = args.text.lower()
    max_r = ws.max_row or 1
    max_c = min(ws.max_column or 1, 200)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, str) and text in v.lower():
                cl = get_column_letter(c)
                print(f"  {cl}{r}: {v!r}")


def main():
    p = argparse.ArgumentParser(description="HeroForge .hfg reader/converter")
    sub = p.add_subparsers(dest="cmd")

    # SRC read
    rd = sub.add_parser("SRC")
    rd.add_argument("sheet")
    rd.add_argument("--rows")
    rd.add_argument("--cols")
    rd.add_argument("--nonempty", action="store_true")
    rd.set_defaults(target="SRC")

    # CHAR read
    rc = sub.add_parser("CHAR")
    rc.add_argument(
        "sheet",
        nargs="?",
        default="ExportSheet",
    )
    rc.add_argument("--rows")
    rc.add_argument("--cols")
    rc.add_argument("--nonempty", action="store_true")
    rc.set_defaults(target="CHAR")

    # lookup
    lu = sub.add_parser("lookup")
    lu.add_argument("--class-ids")
    lu.add_argument("--prc-ids")
    lu.add_argument("--alignment-id")
    lu.add_argument("--deity-id")
    lu.add_argument("--weapon-ids")

    # search
    se = sub.add_parser("search")
    se.add_argument("target", choices=["SRC", "CHAR"])
    se.add_argument("sheet")
    se.add_argument("text")

    # dump
    sub.add_parser("dump")

    # convert
    cv = sub.add_parser("convert")
    cv.add_argument(
        "--race",
        required=True,
        help="Race name (not in .hfg export)",
    )
    cv.add_argument(
        "--out",
        required=True,
        help="Output .char.yaml path",
    )

    args = p.parse_args()
    if args.cmd in ("SRC", "CHAR"):
        cmd_read(args)
    elif args.cmd == "lookup":
        cmd_lookup(args)
    elif args.cmd == "search":
        cmd_search(args)
    elif args.cmd == "dump":
        dump_char()
    elif args.cmd == "convert":
        convert_char(args.race, args.out)
    else:
        p.print_help()


if __name__ == "__main__":
    main()
